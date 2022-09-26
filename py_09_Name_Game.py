

import random
from openpyxl import load_workbook
wb = load_workbook('py_09_Name_Game-TOP-NAMES.xlsx', data_only=True)   #data_only=True -> copying values from excel instead of formulas for the Haveseen cell value
ws = wb.active


# Asking the names of the couples, generating the first letter of the new name of their baby

alphabet = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']

print('')
print('Give me an english couple`s first names and I will tranform them into baby names.')


firstname = input('Husband`s First Name: ')
firstname = firstname.lower()
while True:
    for i in firstname:
        if i not in alphabet:
            print()
            print('Please add a proper english first name, no special characters, no numbers.')
            firstname = input('Husband`s First Name: ')
            break
    else:
        break


secondname = input('Wife`s First Name: ')
secondname = secondname.lower()
while True:
    for i in secondname:         
        if i not in alphabet:
            print()
            print('Please add a proper english second name, no special characters, no numbers.')
            secondname = input('Wife`s First Name: ')
            break
    else:
        break

fn_value = 0
for i in firstname:
    n = alphabet.index(i) + 1   #looking for the position of the first name`s letters in alphabet
    fn_value = fn_value + n     #sum of the position/index of the letters


sn_value = 0
for i in firstname:
    n = alphabet.index(i) + 1
    sn_value = sn_value + n



first_letter_number = fn_value*sn_value

first_letter_number = fn_value*sn_value%26  # 26 letters in the US alphabet

first_letter = alphabet[first_letter_number-1].upper()



# ----- Collecting the correct name(s) from Excel to a list, from the list back to the user -----


boy_list = []   # lists where we collect all the names with the first letter generated from the couples names
girl_list = []

for cellnumber in range(5,105):
    cell_boy = 'B' + str(cellnumber)
    boyname = ws[cell_boy].value
    if boyname[0] == first_letter:
        boy_list = boy_list + [boyname]
    

for cellnumber in range(5,105):
    cell_girl = 'F' + str(cellnumber)
    girlname = ws[cell_girl].value
    if girlname[0] == first_letter:
        girl_list = girl_list + [girlname]


print()
print('If it is a')


if boy_list == []:
    print('boy: 666   / for all the souls on the planet, pleae do not give a life for this "boy" /')
else:
    print('boy: ' + random.choice(boy_list))


if girl_list == []:
    print('girl: Backy  / Have you seen the Pet Sematary? /')
else:
    print('girl: ' + random.choice(girl_list))    

print()